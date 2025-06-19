import os
import json
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

COMPANY_FOLDER = "Fyrirtaeki"
EXPORT_FOLDER = "reports"

def read_json(path):
    with open(path, "r") as f:
        return json.load(f)

def load_users():
    return {u["id"]: u for u in read_json("users.json")}

def ensure_folder(path):
    os.makedirs(path, exist_ok=True)

def format_date(iso):
    return datetime.fromisoformat(iso).strftime("%Y-%m-%d")

def compute_hours(start, end):
    delta = datetime.fromisoformat(end) - datetime.fromisoformat(start)
    return round(delta.total_seconds() / 3600, 2)

def export_company_to_excel(company_name):
    task_totals = defaultdict(float)
    day_shifts = defaultdict(list)
    total_hours = 0.0

    company_path = os.path.join(COMPANY_FOLDER, company_name)
    report_path = os.path.join(EXPORT_FOLDER, f"{company_name}.xlsx")
    ensure_folder(EXPORT_FOLDER)

    users = load_users()

    for file in os.listdir(company_path):
        if not file.endswith(".json"):
            continue

        employee_id = file.replace(".json", "")
        user = users.get(employee_id, {"name": "Unknown", "id": employee_id})
        logs = read_json(os.path.join(company_path, file))

        for log in logs:
            if log["clock_out"] is None:
                continue
            start = log["clock_in"]
            end = log["clock_out"]
            date = format_date(start)
            duration = compute_hours(start, end)
            total_hours += duration
            task_name = log.get("task", "N/A")
            task_totals[task_name] += duration

            day_shifts[date].append([
                user["id"],
                user["name"],
                log.get("location", "N/A"),
                log.get("task", "N/A"),
                start[11:16],
                end[11:16],
                duration
            ])

    wb = Workbook()
    ws = wb.active
    ws.title = "Work Hours"

    bold = Font(bold=True)
    current_row = 1

    for date in sorted(day_shifts.keys()):
        ws.append([f"Date: {date}"])
        ws.cell(row=current_row, column=1).font = bold
        current_row += 1

        headers = ["Employee", "Name", "Location", "Task", "Clock In", "Clock Out", "Hours Worked"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=col).font = bold
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal="center")
        current_row += 1

        day_total = 0.0
        for shift in day_shifts[date]:
            ws.append(shift)
            day_total += shift[-1]
            current_row += 1

        ws.append([""] * 6 + [f"Total: {round(day_total, 2)}"])
        ws.cell(row=current_row, column=7).font = bold
        current_row += 2

    # Overall total
    ws.append([])
    ws.append([""] * 6 + [f"Overall Total Hours: {round(total_hours, 2)}"])
    ws.cell(row=current_row + 1, column=7).font = bold
    current_row += 3

    # Task Summary
    ws.cell(row=current_row + 1, column=1).font = bold
    ws.append(["Task Name", "Total Hours"])
    ws.cell(row=current_row + 2, column=1).font = bold
    ws.cell(row=current_row + 2, column=2).font = bold
    current_row += 3

    for task, hours in sorted(task_totals.items()):
        ws.append([task, round(hours, 2)])
        current_row += 1

    # Autosize columns
    for col in range(1, 8):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].auto_size = True

    wb.save(report_path)
    print(f"✅ Excel report created: {report_path}")

def export_all_companies():
    for company in os.listdir(COMPANY_FOLDER):
        company_path = os.path.join(COMPANY_FOLDER, company)
        if os.path.isdir(company_path):
            export_company_to_excel(company)

if __name__ == "__main__":
    export_all_companies()
