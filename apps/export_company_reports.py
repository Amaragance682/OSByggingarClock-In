import os
import json
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side

from lib.utils import resource_path, load_task_config  # <- import the new loader

COMPANY_FOLDER = resource_path("Database/Fyrirtaeki")
EXPORT_FOLDER  = resource_path("Database/reports")

thin_gray = Border(
    left=Side(style="thin",  color="999999"),
    right=Side(style="thin", color="999999"),
    top=Side(style="thin",   color="999999"),
    bottom=Side(style="thin", color="999999"),
)
thick_black = Border(
    left=Side(style="medium", color="000000"),
    right=Side(style="medium",color="000000"),
    top=Side(style="medium",  color="000000"),
    bottom=Side(style="medium",color="000000"),
)

def read_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def load_users():
    return {u["id"]: u for u in read_json(resource_path("Database/users.json"))}

def ensure_folder(path):
    os.makedirs(path, exist_ok=True)

def format_date(iso):
    dt = datetime.fromisoformat(iso)
    day = dt.day
    suffix = "th" if 11 <= day <= 13 else {1:"st",2:"nd",3:"rd"}.get(day%10,"th")
    return f"{day}{suffix} of {dt.strftime('%B')}"

def compute_hours(start, end):
    delta = datetime.fromisoformat(end) - datetime.fromisoformat(start)
    return round(delta.total_seconds()/3600, 2)

def export_company_to_excel(company_name):
    # — load the enriched task config —
    cfg = load_task_config()

    # define your column headers once, up front
    headers = ["Name","Location","Task","Clock In","Clock Out","Lunch (min)","Paid Hours"]

    task_totals = defaultdict(float)
    day_shifts  = defaultdict(list)
    total_paid  = 0.0

    current_month = datetime.now().strftime("%B")
    company_path   = os.path.join(COMPANY_FOLDER, company_name)
    report_path    = os.path.join(EXPORT_FOLDER, f"{company_name}_{current_month}.xlsx")
    ensure_folder(EXPORT_FOLDER)

    users = load_users()

    # — gather data from each employee log —
    for fn in os.listdir(company_path):
        if not fn.endswith(".json"):
            continue
        eid  = fn[:-5]
        user = users.get(eid, {"id":eid,"name":"Unknown"})
        logs = read_json(os.path.join(company_path, fn))
        for log in logs:
            start = log.get("clock_in")
            end   = log.get("clock_out")
            if not start or not end:
                continue

            dt = datetime.fromisoformat(start).date()
            raw_hours  = compute_hours(start, end)
            lunch_mins = log.get("lunch_minutes", 0)
            paid_hours = round(raw_hours - (lunch_mins / 60), 2)

            total_paid += paid_hours
            task_name = log.get("task", "N/A")
            task_totals[task_name] += paid_hours

            day_shifts[dt].append([
                user["name"],
                log.get("location","N/A"),
                task_name,
                start[11:16],
                end[11:16],
                lunch_mins,
                paid_hours
            ])

    # — build workbook —
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Hours"

    bold        = Font(bold=True, size=12)
    header_font = Font(bold=True)
    section_f   = Font(bold=True, size=14)
    center      = Alignment(horizontal="center")
    row = 1

    if day_shifts:
        for dt in sorted(day_shifts):
            if row > 1: row += 2
            top_row = row

            # date header
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            c = ws.cell(row=row, column=1, value=format_date(dt.isoformat()))
            c.font      = section_f
            c.alignment = Alignment(horizontal="left")
            row += 1

            # column headers (added Lunch column)
            headers = ["Name","Location","Task","Clock In","Clock Out","Lunch (min)","Paid Hours"]
            ws.append(headers)
            for col in range(1, len(headers)+1):
                cell = ws.cell(row=row, column=col)
                cell.font      = header_font
                cell.alignment = center
            row += 1

            # shifts
            for rec in day_shifts[dt]:
                ws.append(rec)
                row += 1

            # day total (sum paid hours)
            day_total = sum(r[-1] for r in day_shifts[dt])
            ws.append([""] * (len(headers)-1) + [f"Total: {round(day_total,2)} hrs"])
            ws.cell(row=row, column=len(headers)).font = bold

            # apply borders
            bottom = row
            for r in range(top_row, bottom+1):
                for c in range(1, len(headers)+1):
                    ws.cell(row=r, column=c).border = thin_gray
            row += 1
    else:
        ws.append([f"No shift data for {company_name} in {current_month}"])
        ws.cell(row=row, column=1).font = section_f
        row += 2

    # overall total
    ws.append([])
    row += 1
    ws.append([""] * (len(headers)-1) + [f"Overall Paid Hours: {round(total_paid,2)} hrs"])
    ws.cell(row=row+1, column=len(headers)).font = section_f
    row += 3

    # — Task Summary (unchanged) —
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1).font = section_f
    row += 1

    ws.append(["Task Name","Total Hours","Completed?"])
    for col in (1,2,3):
        ws.cell(row=row, column=col).font      = header_font
        ws.cell(row=row, column=col).alignment = center
    row += 1

    if task_totals:
        comp_states = {}
        for loc, comps in cfg.items():
            for item in comps.get(company_name, []):
                name = item["name"] if isinstance(item, dict) else item
                comp_states[name] = (item.get("completed", False) if isinstance(item, dict) else False)

        for task, hrs in sorted(task_totals.items()):
            done = comp_states.get(task, False)
            ws.append([task, round(hrs,2), "Yes" if done else "No"])
            row += 1
    else:
        ws.append(["No tasks","0.00","—"])
        row += 1

    # autosize columns
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[letter].width = max_length + 2

    wb.save(report_path)
    print(f"✅ Excel report written to: {report_path}")



def export_all_companies():
    for c in os.listdir(COMPANY_FOLDER):
        if os.path.isdir(os.path.join(COMPANY_FOLDER,c)):
            export_company_to_excel(c)


if __name__=="__main__":
    export_all_companies()
